print('tk...')
from tkinter import filedialog
# import tkinter

import datetime

print('sqlite3...')
import sqlite3

# http://biopython.org/
print('Bio...')
from Bio import SeqIO
from Bio import GenBank       # too ?
#import BioSQL
#from BioSQL import BioSeqDatabase

#import PyQt5


#class RefSchema


class CreateTaxa:
    def __init__(self, db, superkingdom_name, root_taxa_name, NCBI_TaxID  = None, syn=None):
        self.db=db
        self.c = db.cursor()
        self.kingdom = superkingdom_name
        self.root_rank = self._root_rank('superkingdom', superkingdom_name)
        self.root_taxa = self._root_taxa(root_taxa_name, root_taxa_name, NCBI_TaxID)

    def _root_rank (self, name, kingdom):        #  todo: NCBI ? add kingdom ! determine type of rank.
        self.c.execute("INSERT INTO taxa_rank (Name, kingdom )"
                       "               VALUES (?   , ?       )",
                       (                       name, kingdom )   )
        self.root_rank = self.c.lastrowid
        return self.root_rank

    def _root_taxa (self, name, vulgar, NCBI_TaxID  = None, syn=None):
        self.c.execute("INSERT INTO taxa      (Name , vulgar, Id_rank    , NCBI_TaxID  )"
                       "               VALUES ( ?   , ?     , ?          , ?           )",
                       (name , vulgar, self.root_rank, NCBI_TaxID))
        self.root_taxa = self.c.lastrowid
        self.c.execute("INSERT INTO taxa_parents (Id_taxa     , parent     , Id_rank      )"
                       "                  VALUES ( ?          , ?          , ?            )",
                       (self.root_taxa , self.root_taxa, self.root_rank))
        self.synonyms(self.root_taxa, [name, vulgar, NCBI_TaxID])
        if syn: self.synonyms(self.root_taxa, syn)
        return self.root_taxa

    def rank (self, name, parent_rank):
        self.c.execute("INSERT INTO taxa_rank (Name, kingdom     , parent      )"
                       "               VALUES (?   , ?           , ?           )",
                       (                       name, self.kingdom, parent_rank )   )
        return self.c.lastrowid

    def taxa (self, name, vulgar, rank, parent_taxa, NCBI_TaxID  = None, syn=None):

        self.c.execute("INSERT INTO taxa (Name,  vulgar, Id_rank, parent     , NCBI_TaxID )"
                       "          VALUES (?   , ?      , ?      , ?          , ?          )",
                       (name, vulgar , rank   , parent_taxa, NCBI_TaxID))
        Id_taxa=self.c.lastrowid
        self.c.execute("INSERT INTO taxa_parents (Id_taxa, parent, Id_rank      )"
                       "                  VALUES ( ?     , ?     , ?            )",
                                                 (Id_taxa , Id_taxa, rank       ) )
        self.c.execute("INSERT INTO taxa_parents (Id_taxa, parent, Id_rank      )"
                       "                  SELECT  ?      , parent, Id_rank       "
                       "FROM taxa_parents WHERE Id_taxa=?",     (Id_taxa , parent_taxa  ) )

        self.synonyms(Id_taxa, [name, vulgar, NCBI_TaxID])
        if syn: self.synonyms(Id_taxa, syn)
        return Id_taxa

    def synonyms(self, Id_taxa, names):
        if isinstance(names, str):
            names=[names]
        for name in names:
            if not name: continue
            self.c.execute("INSERT INTO taxa_names (Id_taxa, Name      )"
                           "                VALUES ( ?     , ?         )",
                                                   (Id_taxa, name      )  )


def rank_ID(db_cursor, rank_name):
    db_cursor.execute("SELECT Id_rank FROM taxa_rank WHERE Name=?", (rank_name,))
    return db_cursor.fetchone()[0]


def rank_ID_taxa(db_cursor, Id_taxa):
    db_cursor.execute("SELECT Id_rank FROM taxa WHERE Id_taxa=?", (Id_taxa,))
    return db_cursor.fetchone()[0]

#def createBioSQL(newly) -> sqlite3.Connection:
#
#    server = BioSeqDatabase.open_database(driver="sqlite3", db="HEV")
#    try:
#        db = server["HEV"]
#    except KeyError:
#    db = server.new_database("HEV",
#                                 description="For testing GBrowse")
    #db = sqlite3.connect("../data/temp/BioSQL.db")
    #if newly:
       #read_create(db)
       #print('Adding default taxas...')
       #add_def_taxa(db)
       #print('Adding reference schemes...')
       #add_ref_schema(db)
       #db.commit()
#
#   return db
#
#def read_createBioSQL(db):
#    with open("biosqldb-sqlite.sql") as dbcreate:
#        sql_create = dbcreate.read()
#    c = db.cursor()
#    c.executescript(sql_create)


def create(newly) -> sqlite3.Connection:
    db = sqlite3.connect("../data/temp/seq.db")
    if newly:
       read_create(db)
       print('Adding default taxas...')
       add_def_taxa(db)
       print('Adding reference schemes...')
       add_ref_schema(db)
       db.commit()

    return db


def read_create(db):
    with open("create_seq.sql") as dbcreate:
        sql_create = dbcreate.read()
    c = db.cursor()
    c.executescript(sql_create)


def read_country_codes(db):
    with open("country_codes.sql") as dbcreate:
        sql_create = dbcreate.read()
    c = db.cursor()
    c.executescript(sql_create)
    db.commit()


def add_ref_schema(db):
    c = db.cursor()
    for sch in [('Lu'  , 'Lu, Li, 2006'         ),
                ('VR'  , 'Vina-Rodriguez, 2015' ),
                ('ICVT', 'ICVT, 2016'           )   ] :
      c.execute("INSERT INTO ref_schema (schema, name) VALUES (?,     ?)", sch)


def add_def_taxa(db):
    ct = CreateTaxa(db,
                    superkingdom_name = 'Viruses',
                    root_taxa_name    = 'Viridae',
                    NCBI_TaxID        = '10239',
                    syn               = ['Viridae', 'Vira','viruses'])

    realm   = ct.rank(name='Realm', parent_rank=ct.root_rank)  # no rank at NCBI
    Riboviria = ct.taxa(name        = 'Riboviria',
                        vulgar      = 'RNA viruses',
                        rank        = realm,
                        parent_taxa = ct.root_taxa,
                        NCBI_TaxID  = '439488')

    phylum   = ct.rank(name='phylum', parent_rank=realm)
    Negarnaviricota = ct.taxa(name        = 'Negarnaviricota',
                              vulgar      = '',
                              rank        = phylum,
                              parent_taxa = Riboviria,
                              NCBI_TaxID  = '2497569')

    subphylum   = ct.rank(name='subphylum', parent_rank=phylum)
    Polyploviricotina = ct.taxa(name        = 'Polyploviricotina',
                                vulgar      = '',
                                rank        = subphylum,
                                parent_taxa = Negarnaviricota,
                                NCBI_TaxID  = '2497571')

    Class   = ct.rank(name='class', parent_rank=subphylum)
    Ellioviricetes = ct.taxa(name        = 'Ellioviricetes',
                             vulgar      = '',
                             rank        = Class,
                             parent_taxa = Polyploviricotina,
                             NCBI_TaxID  = '2497576')

    order = ct.rank(name='Order', parent_rank=Class)
    Bunyavirales  = ct.taxa(name        = 'Bunyavirales',
                            vulgar      = 'Bunyavirales',
                            rank        = order,
                            parent_taxa = Ellioviricetes,
                            NCBI_TaxID  = '1980410')

    family    = ct.rank(name='family',    parent_rank=order)
    subfamily = ct.rank(name='subfamily', parent_rank=family)

    Arenaviridae     = ct.taxa('Arenaviridae'      , 'Arenaviridae', family, Bunyavirales, '11617')
    Hantaviridae     = ct.taxa('Hantaviridae'      , 'Hantaviridae', family, Bunyavirales, '1980413')
    Nairoviridae     = ct.taxa('Nairoviridae'      , 'Nairoviridae', family, Bunyavirales, '1980415')
    Peribunyaviridae = ct.taxa('Peribunyaviridae'  , 'Peribunyaviridae', family, Bunyavirales, '1980416')
    Phenuiviridae    = ct.taxa('Phenuiviridae'     , 'Phenuiviridae', family, Bunyavirales, '1980418')

    tHEV   = ct.taxa('Hepeviridae', 'HEV'  , family, Riboviria, '291484',
                   syn=['2021911', '1009842', '172851', '2021912', '2021913', '2021914', '1216472', '996468',
                       '1638959', '1638960', '1674928', '1530451', '1328106', '1229326', '879095', '301242'])

    # 172851 Avian hepatitis E virus, 2021912 Barns Ness breadcrumb sponge hepe-like virus 2,
    # 2021913, Barns Ness breadcrumb sponge hepe-like virus 3, 2021914 Barns Ness breadcrumb sponge hepe-like virus 4
    # 1216472, Bat hepevirus;  996468 Hepatitis E virus rat/USA/2003; 1638959 Mystacina/New Zealand/2013/3
    # 1638960 Mystacina/New Zealand/2013;  1674928 seal/AAUST73/BR/2012 ; 1530451 Fesavirus 2; 1328106 Fox
    # 1229326 Hepelivirus; 879095 rat/R68/DEU/2009; 301242 Big liver and spleen disease virus

    genus  = ct.rank(name='genus',  parent_rank=subfamily)
    Orthobunyavirus = ct.taxa('Orthobunyavirus', 'Bunyavirus', genus, Peribunyaviridae, '11572', syn=['Bunyaviruses'])  # syn ?
    Orthonairovirus_genus = ct.taxa('Orthonairovirus', 'Nairovirus', genus, Nairoviridae, '1980517')
    Phlebovirus = ct.taxa('Phlebovirus',     'Phlebovirus', genus, Phenuiviridae, '11584')

    tOrthHEV= ct.taxa('Orthohepevirus', 'Orthohepevirus'  , genus, tHEV, '1678141', syn=['12461', 'Hepatitis E virus','186677', 'Hepevirus'])
    tPisci  = ct.taxa('Piscihepevirus', 'Piscihepevirus'  , genus, tHEV, '1678142')

    species  = ct.rank('species', parent_rank=genus)

    tBySp    = ct.taxa('Bunyamwera orthobunyavirus', 'Bunyamwera orthobunyavirus'  , species, Orthobunyavirus, '1933179')
    rSubSpecie = ct.rank('subspecie', species)
    tByVr    = ct.taxa('Bunyamwera virus', 'Bunyamwera serogroup'  , rSubSpecie, tBySp, '35304', syn=['Bunyamwera virus group', 'Bunyamwera bunyavirus group'])
    tBtVr    = ct.taxa('Batai virus', 'Batai'  , rSubSpecie, tBySp, '80942' )

    # Serogroup Crimean-Congo hemorrhagic fever
    CCHFV_Sp    = ct.taxa('Crimean-Congo hemorrhagic fever orthonairovirus', 'CCHFV', species, Orthonairovirus_genus, '1980519')
    HAZV_Sp     = ct.taxa('Hazara orthonairovirus',   'HAZV',     species, Orthonairovirus_genus, '1980522')
    # Tofla (TFLV)

    # Serogroup Dera Ghazi Khan
    DGKV_Sp = ct.taxa('Dera Ghazi Khan orthonairovirus', 'DGKV', species, Orthonairovirus_genus, '1980520')
    # Abu Hammad  (AHV)
    # Abu Mina  (AMV)

    # Serogroup Hughes
    HUGV_Sp   = ct.taxa('Hughes orthonairovirus',   'HUGV',   species, Orthonairovirus_genus, '1980523')
    # Caspiy  (CASV)
    # Farallon (FARV)
    # Punta salinas  (PSV)
    # Raza  (RAZAV)
    # Soldado  (SOLV)
    # Zirqa (ZIRV)

    # Serogroup Sakhalin
    SAKV_Sp = ct.taxa('Sakhalin orthonairovirus', 'SAKV', species, Orthonairovirus_genus, '1980528')
    # Avalon  (AVAV)
    # Paramushir  (PMRV)
    # Tilamook  (TILV)

    # Serogroup Nairobi sheep disease
    NSDV_Sp    = ct.taxa('Nairobi sheep disease orthonairovirus', 'NSDV', species, Orthonairovirus_genus, '1980526')
    DUGV_Sp    = ct.taxa('Dugbe orthonairovirus',                 'DUGV', species, Orthonairovirus_genus, '1980521')
    # Ganjam  (GANV)
    # Kupe (KUPV)

    # Serogroup Qalyub
    QYBV_Sp   = ct.taxa('Qalyub orthonairovirus',   'QYBV',   species, Orthonairovirus_genus, '1980527')
    CHIMV_Sp  = ct.taxa('Chim orthonairovirus',     'CHIMV',  species, Orthonairovirus_genus, '2170062')
    # Geran  (GERV)
    # Bandia  (BDAV)

    # Serogroup Thiafora
    TFAV_Sp = ct.taxa('Thiafora orthonairovirus', 'TFAV', species, Orthonairovirus_genus, '1980529')
    # Erve (ERVV)

    # Serogroup Issyk-kul
    # Issyk-kul (ISKV)
    # Gossas (GOSV)
    # Uzun Agach (UZAV)

    # Serogroup Kasokero
    KKOV_Sp = ct.taxa('Kasokero orthonairovirus', 'KKOV', species, Orthonairovirus_genus, '1980524')
    # Leopards Hill  (LPHV)
    # Yogue (YOGV)

    # Serogroup Burana
    BURV_Sp = ct.taxa('Burana orthonairovirus', 'BURV', species, Orthonairovirus_genus, '1980518')  # not official?
    TDYV_Sp = ct.taxa('Tamdy orthonairovirus',  'TDYV', species, Orthonairovirus_genus, '2170063')

    Artashat_Sp = ct.taxa('Artashat orthonairovirus', 'Artashat', species, Orthonairovirus_genus, '2170061')
    Keterah_Sp  = ct.taxa('Keterah orthonairovirus',  'Keterah',  species, Orthonairovirus_genus, '1980525')
    Qalyub_Sp   = ct.taxa('Qalyub orthonairovirus',   'Qalyub',   species, Orthonairovirus_genus, '1980527')
    # Estero Real orthonairovirus : https://talk.ictvonline.org//taxonomy/p/taxonomy-history?taxnode_id=201850107
    # Estero Real orthobunyavirus : https://www.ncbi.nlm.nih.gov/Taxonomy/Browser/wwwtax.cgi?name=Estero+Real+virus

    tOrthSpcA= ct.taxa('Orthohepevirus A', 'Orthohepevirus A'  , species, tOrthHEV, '1678143', syn=[ 'Swine hepatitis E virus', '63421']) # ?? syn??
    ct.synonyms(tOrthSpcA, '12461')
    tOrthSpcB= ct.taxa('Orthohepevirus B', 'Orthohepevirus B'  , species, tOrthHEV, '1678144') # NCBI_ID tentative !!
    tOrthSpcC= ct.taxa('Orthohepevirus C', 'Orthohepevirus C'  , species, tOrthHEV, '1678145', syn=['879096', '1414752']) # rat/R63/DEU/2009, Mink
    tOrthSpcD= ct.taxa('Orthohepevirus D', 'Orthohepevirus D'  , species, tOrthHEV, '1678146')


    tPisciSpcA=ct.taxa('Piscihepevirus A', 'Piscihepevirus A'  , species, tPisci,'1678146', syn=['1016879']) # Cutthroat trout virus

    rGenotype = ct.rank('genotype', species)
    g1  = ct.taxa('1', 'HEV-g1'  , rGenotype, tOrthSpcA, '185579', syn=['I', 'GI', 'G1', 'One'] )
    g2  = ct.taxa('2', 'HEV-g2'  , rGenotype, tOrthSpcA, syn=['II'] )
    g3  = ct.taxa('3', 'HEV-g3'  , rGenotype, tOrthSpcA, '509628', syn=['G3', 'III', 'Gt3', 'g3', 'HEV-3', 'third', 'GIII'])   # , 'Hepatitis E virus type 3'
    g4  = ct.taxa('4', 'HEV-g4'  , rGenotype, tOrthSpcA, '185580',
                  syn=['IV', '689698', '689699', '689700', '689701', '689702', '689703', '4(IV)']) # Hu/03858/HKG/2009, Hu/07598/HKG/2006
    g5  = ct.taxa('5', 'HEV-g5'  , rGenotype, tOrthSpcA, syn=['V'] )
    g6  = ct.taxa('6', 'HEV-g6'  , rGenotype, tOrthSpcA, syn=['VI']  )
    g7  = ct.taxa('7', 'HEV-g7'  , rGenotype, tOrthSpcA, syn=['VII', 'HEV-7' ] )

    gC1 = ct.taxa('C1', 'HEV-C1' , rGenotype, tOrthSpcC )
    gC2 = ct.taxa('C2', 'HEV-C2' , rGenotype, tOrthSpcC, syn=['1213422', 'Ferret hepatitis E virus'] )

    rmc  = ct.rank('major clade', rGenotype)
    maI  = ct.taxa('I'  , 'HEV-g3-I'     , rmc, g3)
    maII = ct.taxa('II' , 'HEV-g3-II'    , rmc, g3)
    Rab  = ct.taxa('3ra', 'HEV-g3-rabbit', rmc, g3)

    rgr  = ct.rank('group', rmc)
    grchi  = ct.taxa('3chi'  , 'HEV-g3chi'     , rgr, maI)
    grjab  = ct.taxa('3jab'  , 'HEV-g3jab'     , rgr, maI)
    grfeg  = ct.taxa('3feg'  , 'HEV-g3feg'     , rgr, maII)
    grRab  = ct.taxa('3rab'  , 'HEV-g3rabbit'  , rgr, Rab, syn=['Rab']  )  # temporal???

    rsubt = ct.rank('subtype', rgr)

    g1a   = ct.taxa('1a', 'HEV-g1a'  , rsubt, g1, syn=['a', 'Ia', 'IA'])
    g1b   = ct.taxa('1b', 'HEV-g1b'  , rsubt, g1)
    g1c   = ct.taxa('1c', 'HEV-g1c'  , rsubt, g1)
    g1d   = ct.taxa('1d', 'HEV-g1d'  , rsubt, g1, syn=['d'])
    g1e   = ct.taxa('1e', 'HEV-g1e'  , rsubt, g1)
    g1f   = ct.taxa('1f', 'HEV-g1f'  , rsubt, g1)
    g1g   = ct.taxa('1g', 'HEV-g1g'  , rsubt, g1)
    g1h   = ct.taxa('1h', 'HEV-g1h'  , rsubt, g1)
    g1i   = ct.taxa('1i', 'HEV-g1i'  , rsubt, g1)
    g1j   = ct.taxa('1j', 'HEV-g1j'  , rsubt, g1)
    g1k   = ct.taxa('1k', 'HEV-g1k'  , rsubt, g1)

    g2a   = ct.taxa('2a', 'HEV-g2a'  , rsubt, g2)

    g3a   = ct.taxa('3a', 'HEV-g3a'  , rsubt, grjab, syn=['a'])
    g3b   = ct.taxa('3b', 'HEV-g3b'  , rsubt, grjab)
    g3c   = ct.taxa('3c', 'HEV-g3c'  , rsubt, grchi, syn=['c', 'G3c'])
    g3d   = ct.taxa('3d', 'HEV-g3d'  , rsubt, g3, syn=['d'])
    g3e   = ct.taxa('3e', 'HEV-g3e'  , rsubt, grfeg, syn=['e', 'g3e', 'G3E', '3E'])
    g3ef  = ct.taxa('3ef', 'HEV-g3ef', rsubt, grfeg )
    g3f   = ct.taxa('3f', 'HEV-g3f'  , rsubt, grfeg, syn=['f', 'g3f', 'G3F', '3F', '515413', '515412']) # human/3f/Fr-27/France/2006, Fr-26/France/2006
    g3g   = ct.taxa('3g', 'HEV-g3g'  , rsubt, grfeg)
    g3h   = ct.taxa('3h', 'HEV-g3h'  , rsubt, grchi, syn=['h', 'g3h', 'G3H', '3H'])
    g3i   = ct.taxa('3i', 'HEV-g3i'  , rsubt, grchi)
    g3j   = ct.taxa('3j', 'HEV-g3j'  , rsubt, grjab)
    g3k   = ct.taxa('3k', 'HEV-g3k'  , rsubt, g3)
    g3l   = ct.taxa('3l', 'HEV-g3l'  , rsubt, g3)

    g4a   = ct.taxa('4a', 'HEV-g4a'  , rsubt, g4)
    g4b   = ct.taxa('4b', 'HEV-g4b'  , rsubt, g4)
    g4c   = ct.taxa('4c', 'HEV-g4c'  , rsubt, g4)
    g4d   = ct.taxa('4d', 'HEV-g4d'  , rsubt, g4, syn=['d'])
    g4e   = ct.taxa('4e', 'HEV-g4e'  , rsubt, g4)
    g4f   = ct.taxa('4f', 'HEV-g4f'  , rsubt, g4)
    g4g   = ct.taxa('4g', 'HEV-g4g'  , rsubt, g4)
    g4h   = ct.taxa('4h', 'HEV-g4h'  , rsubt, g4, syn=['h'])
    g4i   = ct.taxa('4i', 'HEV-g4i'  , rsubt, g4)
    g4j   = ct.taxa('4j', 'HEV-g4j'  , rsubt, g4)
    g4k   = ct.taxa('4k', 'HEV-g4k'  , rsubt, g4)

    db.commit()


def build_ref_pos(Seq, beg, end, Al_len):
    sr=0
    ref = [sr]*beg
    for b in Seq:
        if (b != '-'): sr+=1
        ref.append(sr)
    ref += [sr]*(Al_len - end-1)
    return ref


def parse_full_fasta_Align(db, ref_seq = None, file_name=None):
    """Will parse an alignment and insert it into the tables:
        files: the file path ,
        align:
        seq: because assume this sequences are all complete original sequences (but may be a partial sequence from
             some isolate ! )

        """
    if not file_name:
        file_name = filedialog.askopenfilename(filetypes=(("fasta aligment", "*.fas"), ("All files", "*.*")),
                                               defaultextension='fas',
                                               title='Select Master Alignment')
        if not file_name:
            return

    # self.refSeq.clear()
    print(file_name)

    c = db.cursor()
    c.execute("INSERT INTO seq_file (path, format) VALUES (?, 'fasta')", (file_name,))

    Id_file = c.lastrowid
    c.execute("INSERT INTO align (Id_file, Name,      Ref     ) "
              "           VALUES (?,       ?,         ?       )",
                                 (Id_file, file_name, ref_seq )      )

    Id_align = c.lastrowid
    max_len = 0
    for seq_record in SeqIO.parse(file_name, "fasta"):
        # print(seq_record.id, len(seq_record) )
        seq = str(seq_record.seq)
        if ref_seq is None:      # set first seq as reference
            ref_seq = seq_record.id
            al_ref_seq = seq
        else:
            if ref_seq == seq_record.id:
                al_ref_seq = seq
        ln = len( seq)
        if max_len < ln: max_len = ln
        seq_beg = 0
        seq_end = ln - 1
        while seq_beg < ln:
            if  seq[seq_beg] == '-':
                seq_beg += 1
            else:
                break
        while seq_end > seq_beg:
            if seq[seq_end] == '-':
                seq_end -= 1
            else:
                break

        seq = str( seq[seq_beg: seq_end + 1])

        c.execute("SELECT Id_seq FROM seq WHERE seq.Name=?", (str(seq_record.id),)) # (record.organism,))
        Id_part = c.fetchone()
        Id_part = Id_part[0] if Id_part else None
        if not Id_part:
            exp_seq = seq.replace('-','')  #''.join([base for base in seq if base != '-'])

            c.execute("INSERT INTO seq (Name,               Seq,     Len  ) "
                  "         VALUES (?,                  ?,       ?    )",
                                   (str(seq_record.id), exp_seq, len(exp_seq))    )
            Id_part = c.lastrowid

        c.execute("INSERT INTO aligned_seq (Id_align, Id_part, Seq,      pbeg,     pend  ) "
                  "                 VALUES (?,        ?,       ?,        ?,       ?    )",
                                           (Id_align, Id_part, seq, seq_beg, seq_end )    )

    c.execute("UPDATE align SET Al_len = ?, Ref=?    WHERE Id_align=?",
                            (   max_len   , ref_seq,        Id_align ) )
    db.commit()
    return Id_align , build_ref_pos(al_ref_seq,0,len(al_ref_seq)-1, max_len)


def ref_pos(sdb, ID_align, seq_name=None):
    c = sdb.cursor()
    c.execute("SELECT Al_len, Ref FROM align WHERE Id_align=? ", (ID_align,  ))
    Al_len, Ref = c.fetchone()
    if not seq_name: seq_name = Ref
    c.execute("SELECT aligned_seq.Seq, pbeg, pend FROM aligned_seq, Seq "
              "ON Id_part=Id_seq WHERE Name=? AND Id_align=?",
                                     (seq_name,     ID_align ))
    Seq, beg, end = c.fetchone()
    return build_ref_pos(Seq, beg, end, Al_len)


def parse_row(db, row, col):
    success = True
    MEGA_name = row[col['MEGA name']].value
    genotype  = row[col['genotype']].value
    subtype   = row[col['subtype']].value
    group     = row[col['group']].value
    Str_name  = row[col['Str.name']].value
    Isolate   = row[col['Isolate']].value
    # Country  = row[col['Country'     ]].value
    Country_cod3=row[col['Country cod']].value
    Region     =row[col['region']].value         # 'H '        todo: deduced
    Region_full=row[col['region full']].value
    Host       =row[col['Host']].value
    Source     =row[col['Source']].value
    Year       =row[col['Y']].value
    Month      =row[col['M']].value
    Day        =row[col['D']].value
    Institut   =row[col['Inst']].value
    Reference  =row[col['reference']].value
    Lu_Li      =row[col['Lu, Li']].value
    CG         =row[col['CG']].value

    c = sdb.cursor()
    taxa = subtype if subtype else group if group else genotype
    c.execute("SELECT Id_taxa FROM taxa WHERE taxa.Name=?", (taxa, ))   # ?? Name UNIQUE ??
    Id_taxa = c.fetchone()
    Id_taxa = Id_taxa[0] if Id_taxa else None

    if not Isolate: Isolate   = Str_name

    Id_strain  = None
    Id_isolate = None
    Id_seq     = None
    strain_Name = None
    isolate_Name = None

    # let assume the normal situation where seq name exist, and lets examine strain and isolate.
    c.execute("""
                select  seq.Id_seq, 
                        strain.Id_strain, strain.Name, 
	                    isolate.Id_isolate, isolate.Name 

                from      seq 
                left join isolate_seq USING(Id_seq)
                left join strain      USING(Id_strain)  
                left join isolate     USING(Id_isolate)
	
                where seq.Name = ?
                """
              , (            MEGA_name,        ))

    select = c.fetchone()
    if select:
        Id_seq, Id_strain, strain_Name, Id_isolate, isolate_Name = select
    else:
        # there are not seq, but still could be strain and isolate
        print('Unregistered seq ', MEGA_name)

    if not Id_strain:
        c.execute("select Id_strain from strain where Name = ?", (Str_name,))
        Id_strain = c.fetchone()
        if Id_strain:
            Id_strain = Id_strain[0]
            c.execute("select Id_isolate from isolate where Id_strain = ? and Name=?"
                                                              , (Id_strain, Isolate))
            Id_isolate = c.fetchone()
            Id_isolate = Id_isolate[0] if Id_isolate else None

    if Id_strain:
        if not strain_Name or Str_name != strain_Name:
            print('Renaming? strain  of seq ', MEGA_name, ' from ', strain_Name, ' to ', Str_name)
            if not Str_name:
                Str_name = strain_Name  # the same with the other fields !!!!!
        st = strain_Name if strain_Name else Str_name
        # update strain
        c.execute("UPDATE strain SET Name=?  , Id_taxa=?, year=?, host=?, source=?, country_iso3=?  where Id_strain=? "
                                  , (st      , Id_taxa  , Year ,  Host  , Source  , Country_cod3,         Id_strain    ) )
    else:
        # create strain
        print('New Strain:', Str_name)
        c.execute("INSERT INTO strain (Name    , Id_taxa, host, source, year, country_iso3) "
                  "     VALUES        (?       , ?      , ?   , ?     , ?   , ?           ) ",
                                      (Str_name, Id_taxa, Host, Source, Year, Country_cod3))
        Id_strain = c.lastrowid

    if Id_isolate:
        if not isolate_Name or Isolate != isolate_Name:
            print('Renaming? isolate of seq ', MEGA_name, ' from ', isolate_Name, ' to ', Isolate)
            if not Isolate:
                Isolate = isolate_Name     # the same with the other fields !!!!!

        iso = isolate_Name if isolate_Name else Isolate
        c.execute("UPDATE isolate SET Name=?, Id_strain=?, year=?, month=?, day=?, host=?, source=?, institution=?, country_iso3=?, region=?, region_full=?   where Id_isolate=? "
                                  ,    (iso, Id_strain   , Year ,  Month , Day  , Host  , Source  , Institut     , Country_cod3  , Region  , Region_full,          Id_isolate ))
    else:
        c.execute(
            "INSERT INTO isolate (Name   , Id_strain, Year ,  Month , Day, host, source, institution, country_iso3, region, region_full ) "
            "             VALUES (?      , ?        , ?    , ?      , ?  , ?   , ?     , ?          , ?           , ?     , ?           ) "
                               , (Isolate, Id_strain, Year ,  Month , Day, Host, Source, Institut   , Country_cod3, Region, Region_full ))
        Id_isolate = c.lastrowid

    # create isolate_seq and strain_isolate
    c.execute("INSERT INTO isolate_seq (authority, Id_isolate, Id_seq, Id_strain, Id_taxa, Name   , Year ,  Month , Day, host, source, country_iso3, region, region_full ) "
              "VALUES                  ('AV'     , ?         ,?      , ?        , ?      , ?      , ?    , ?      , ?  , ?   , ?     , ?           , ?     , ?           ) "
              ,                        (           Id_isolate, Id_seq, Id_strain, Id_taxa, Isolate, Year ,  Month , Day, Host, Source, Country_cod3, Region, Region_full ))
    Id_isolate_seq= c.lastrowid

    c.execute("INSERT INTO strain_isolate (authority, Id_strain, Name    , Id_isolate_seq) "
              "     VALUES                ('AV'     , ?        , ?       , ?             ) ",
                                          (           Id_strain, Str_name, Id_isolate_seq))

    # todo: revise this. Is general??
    c.execute("SELECT Id_algseq FROM aligned_seq WHERE Id_part=?", (Id_seq,))
    Id_algseq= c.fetchone()
    Id_algseq = Id_algseq[0] if Id_algseq else None

    if Id_algseq is None:
        #success = abnormal_row(c, row)
        c.execute("INSERT INTO pending_seq (Id_taxa, Name,      Id_seq, Id_isolate) VALUES (?,?,?,?)",
                                           (Id_taxa, MEGA_name, Id_seq, Id_isolate))

        # print("Abnormal row !!!!! ", MEGA_name, subtype, Str_name,
        #   "-------> Taxa:{0}, Alseq:{1}, Seq:{2}".format(Id_taxa, Id_algseq, Id_seq))
        success = False
    else:
        c.execute("INSERT INTO classified_seq (Id_taxa, Id_algseq) VALUES (?,?) "
                                            , (Id_taxa, Id_algseq)                 )
        # if c.rowcount == 0: return abnormal_row(c, row)

    rfs = []
    if Lu_Li         : rfs.append( 'Lu' )
    if Reference     : rfs.append( 'VR' )
    if Reference=='R': rfs.append( 'ICVT')
    # print (rfs)
    for rf in rfs:
       # print ('Add scheme: ',Id_taxa ,     rf                ,     Id_seq,   MEGA_name    )
       c.execute("INSERT INTO ref_seq (Id_taxa,                                    Id_ref_schema     , Id_seq, name   ) "
               "               VALUES (?      , (SELECT Id_ref_schema FROM ref_schema WHERE schema=?), ?     , ?      ) "
                                    , (Id_taxa,                                                  rf  , Id_seq, MEGA_name    ))

    # db.commit()
    return success


def parse_HEV_xlsm(db, file_name=None):
    print('openpyxl...')
    import openpyxl

    if not file_name:
        file_name = filedialog.askopenfilename(filetypes=(("Excel files", "*.xlsm"), ("All files", "*.*")),
                                               defaultextension='fas',
                                               title='Select HEV isolate subtyping deta')
        if not file_name:
            return

    # self.refSeq.clear()
    print(file_name)

    wb = openpyxl.load_workbook(file_name)
    print(wb.sheetnames)

    ws = wb['Seq-class']   # 2 ('Seq-class')
    first = True
    error = False
    col=dict()
    for r in ws.iter_rows() :
        if first:
            for c in range(25):    # make a dic with the first 25 columns headers
                col[r[c].value]=c
            first = False
        else:
            error |= parse_row(db,r,col)
    if error:
        print('There were errors during parsing the Excel file !!!!!!!!!!!!!!')

    db.commit()


def clean_parsed_Excel(db):
    c = db.cursor()
    c.execute("DELETE FROM strain")   # ??
    c.execute("DELETE FROM isolate")   # ??
    c.execute("DELETE FROM isolate_seq")   # ??
    c.execute("DELETE FROM pending_seq")   # ??
    c.execute("DELETE FROM classified_seq")   # ??
    db.commit()


def parseGB(db, GB_flat_file=None):
    '''
    Load and parse a GenBank sequence flat file
    :return:
    '''
    if not GB_flat_file:
        GB_flat_file = filedialog.askopenfilename(filetypes=(("Seq flat GB", "*.gb"), ("All files", "*.*") ),
                                                      title='Parse the GenBank sequences in flat format')
    print(GB_flat_file)

    c = db.cursor()
    Id_genotype = rank_ID(c, 'genotype')
    Id_subtype = rank_ID(c, 'subtype')

    # no real need to save this info
    c.execute("INSERT INTO seq_file (path, format) VALUES (?, 'GB_flat')", (GB_flat_file,))
    Id_file = c.lastrowid
    with open(GB_flat_file) as GB_flat:
      prev_sub_auth = None
      prev_sub_jour = None
      prev_sub_ID   = None
      prev_sub_ref  = None
      for record in GenBank.parse(GB_flat):

        # record.locus  : locus - The name specified after the LOCUS keyword in the GenBank
        # record. This may be the accession number, or a clone id or something else.

        Name     = record.accession[0] if record.accession else record.locus
        strain   = None     # record.locus ??
        isolate  = None
        host     = None
        country  = None      # colection contry, NO author country
        region   = None      # colection region, NO author country
        collection_date = None
        year     = None
        month    = None
        day      = None
        source   = None
        genotype = None
        subtype  = None
        NCBI_TaxID = None

        # let assume the normal situation where we are inserting a new seq
        try:
            c.execute("INSERT INTO seq (Name,     Seq             ,   Len               )"
                      "         VALUES (?,          ?             ,   ?                 )",
                                       (Name, str(record.sequence), len(record.sequence)))
            Id_seq = c.lastrowid

        except sqlite3.IntegrityError:
            print('Atempt to duplicate seq ', Name)
            continue

        #  http://biopython.org/DIST/docs/api/Bio.GenBank.Record-pysrc.html#Feature
        for feature in record.features:
            if feature.key == 'source':
                # http://biopython.org/DIST/docs/api/Bio.GenBank.Record.Qualifier-class.html
                for q in feature.qualifiers:

                    if q.key                                    == '/strain=':
                        strain = q.value[1:-1].strip()

                    elif q.key                                  == '/isolate=':
                        isolate = q.value[1:-1].strip()

                    elif q.key                                  == '/country=':
                        country = q.value[1:-1].split(':')
                        if len(country) > 1:
                            region = country[1].strip()  # ok?
                        country = country[0].strip()

                    elif q.key                                  == '/collection_date=':
                        collection_date = q.value[1:-1].strip()
                        year, month, day = parse_date(collection_date)

                    elif q.key == '/source=' or q.key           == '/isolation_source=':
                        source = q.value[1:-1].strip()

                    elif q.key                                  == '/host=':
                        host = q.value[1:-1]

                    elif q.key                                  == '/db_xref=':      # example value= taxon:509628
                         val = q.value[1:-1]
                         m = val.split(':')
                         if m[0].startswith                         ('taxon'):
                             NCBI_TaxID = m[1].strip()

                    elif q.key                                  == '/note=':
                        val = q.value[1:-1]
                        genotype, subtype = parse_GB_note(val)


        authors = None
        unp_aut = None
        references = []
        title = None
        sub_date = None
        institution = None
        Id_submission = None
        p_authors = None
        p_country = None
        p_journal = None

        for rf in record.references:

            if rf.title == 'Direct Submission':
                if prev_sub_auth == rf.authors and prev_sub_jour == rf.journal:
                    Id_submission = prev_sub_ID
                    references = prev_sub_ref
                    break
                else:
                    p_authors = rf.authors
                    p_journal = rf.journal
                    institution = p_journal[24:]

            elif rf.journal == 'Unpublished':
                title = rf.journal
                unp_aut = rf.authors

            else:
                references.append(rf)

        if not Id_submission:    # if not like the previous ...
                                 # look for this authors in existing submissions in DB
            for ID, tl in c.execute("SELECT Id_submission, title  FROM submission WHERE  authors=? ", (p_authors, )):
                if tl == title:           # with the same title will be the same submission
                    Id_submission = ID
                    break
                elif title:
                    continue
                else:
                    for rf in references:   # the submission could uses the title of one of the references
                        if tl == rf.title:
                            title = rf.title
                            Id_submission = ID
                            break

            if not Id_submission and p_journal:   # is a new submission that we can built
                sub_date = p_journal[11:23]       # todo: parse ???
                p_country = institution.split(',')[-1]
                if not title:
                   title = references[0].title if references else None         # lets take the title of the first reference
                c.execute(
                  "INSERT INTO submission ( title, sub_date, authors  , country_iso3                                )"
                  "                VALUES ( ?    ,    ?    , ?        , (SELECT iso3 FROM countries WHERE name=? )  )",
                                          ( title, sub_date, p_authors, p_country                                )  )
                Id_submission = c.lastrowid
                prev_sub_auth = p_authors          # now this will be the prev sub
                prev_sub_jour = p_journal
                prev_sub_ID   = Id_submission
                prev_sub_ref  = references

        for rf in references:   # very often this is empty. Add to every seq referenced
            try:
                c.execute(
                "INSERT INTO reference ( title   ,    authors,    journal   , medline_id,    number,    pubmed_id,    remark) "
                "     VALUES           ( ?       ,    ?      ,    ?      ,    ?         ,    ?     ,    ?        ,    ?     )",
                                       ( rf.title, rf.authors, rf.journal, rf.medline_id, rf.number, rf.pubmed_id, rf.remark))
                reference_id = c.lastrowid

            except sqlite3.IntegrityError:
                c.execute("SELECT reference_id FROM reference WHERE title=? AND authors=? AND journal=?",
                                                                  (rf.title,    rf.authors,   rf.journal))
                reference_id = c.fetchone()[0]
                # reference_id = reference_id[0] if reference_id else reference_id

            c.execute(
                "INSERT INTO reference_to_seq ( location, reference_id , Id_seq  ) "
                "                      VALUES ( ?       , ?            , ?       )",
                                              (rf.bases, reference_id  , Id_seq ))

        Id_taxa = find_ID_Taxa(c, NCBI_TaxID, genotype, subtype, Id_genotype, Id_subtype)
        # Id_rank = rank_ID_taxa(c, Id_taxa)

        c.execute("SELECT iso3 FROM countries LEFT JOIN country_names USING (iso3) "
                  "WHERE countries.name=? or countries.full_name=? OR country_names.name=?", (country, country, country))
        country_iso3 = c.fetchone()
        if not country_iso3:
            if country: print('Not found country: ', country)
            elif region: print('Seq ', Name, ' have no country, but region:', region)
            country_iso3 = None
        else:
            country_iso3 = country_iso3[0]


        host_t, source_t = parse_source(host, source)
        # print('host_t, source_t = parse_source(host, source): ', host_t, source_t, host, source)
        # todo: parse location. Is unique?

        isolate, strain = reuse_GBdefinition_to_find_strain_isolate(record.definition, isolate, strain)
        if not strain: strain = isolate
        if not isolate: isolate = strain
        if not strain:
            strain = Name
            isolate = record.locus

        # there could be strain and isolate, unfortunately most of the time with the same name
        Id_strain = None
        Id_isolate = None

        c.execute("select Id_strain from strain where Name = ?", (strain,))
        Id_strain = c.fetchone()

        if not Id_strain:
            #  the normal situation: a new strain
            c.execute("INSERT INTO strain (Name  , Id_taxa, host  , source  , year, country_iso3) "
                      "     VALUES        (?     , ?      , ?     , ?       , ?   , ?           ) ",
                                          (strain, Id_taxa, host_t, source_t, year, country_iso3))
            Id_strain = c.lastrowid
        else:
            # TENTATIVELY we considere it is the same strain, just these strain have seq already
            # print('Existing Strain:', strain, Id_strain )   # todo: check this  !!!
            # c.execute("UPDATE strain (Name) VALUES (?) ", (strain,))  # todo: update here !!!!!!!!
            # todo: for example - a more precise genotyping
            # Id_strain = c.lastrowid
            # print('     new Id_strain:', Id_strain, )
            Id_strain = Id_strain[0]

            # select other fields to update if possible
            c.execute("select Id_isolate from isolate where Id_strain = ? and Name=?"
                                  , (                       Id_strain,     isolate   ) )
            Id_isolate = c.fetchone()
            if Id_isolate:
                #  a new seq for that isolate of that strain
                # print('Existing isolate:', isolate, Id_isolate )
                # c.execute("UPDATE  isolate (Name) VALUES (?) ", ( isolate,))  # todo: update here !!!!!!!!
                # Id_isolate = c.lastrowid
                Id_isolate = Id_isolate[0]

        if not Id_isolate:
            #  normal situation: a new isolate for that strain
            c.execute(
                "INSERT INTO isolate (Name   , Id_strain, col_date       , Year ,  Month , Day, host  , source  , authors  , institution, country_iso3, region_full ) "
                "             VALUES (?      , ?        , ?              , ?    ,  ?     , ?  , ?    , ?        , ?        , ?          , ?           , ?           ) "
                                   , (isolate, Id_strain, collection_date, year ,  month , day, host_t, source_t, p_authors, institution, country_iso3, region))
            Id_isolate = c.lastrowid

        # create isolate_seq and strain_isolate
        c.execute("INSERT INTO isolate_seq (authority, Id_isolate, Id_seq, Id_submission, Id_strain, Id_taxa, Name   , col_date       , year ,  month , day, host  , source  ,  host_ori, source_ori, country_iso3, region_full) "
                  "VALUES                  ('GB'     , ?         ,?      , ?            , ?        , ?      , ?      , ?              , ?    ,  ?     , ?  , ?     , ?       ,  ?       , ?         , ?           , ?           ) "
                  ,                        (           Id_isolate, Id_seq, Id_submission, Id_strain, Id_taxa, isolate, collection_date, year ,  month , day, host_t, source_t,  host    , source    , country_iso3, region      ))
        Id_isolate_seq= c.lastrowid

        c.execute("INSERT INTO strain_isolate (authority, Id_strain, Name  , Id_isolate_seq) "
                  "     VALUES                ('GB'     , ?        , ?     , ?             ) ",
                                              (           Id_strain, strain, Id_isolate_seq) )

        c.execute("INSERT INTO pending_seq (Name, Id_taxa, description      , Id_isolate, Id_seq) "
                  "     VALUES             (?   , ?      , ?                , ?         ,      ?) ",
                                           (Name, Id_taxa, record.definition, Id_isolate, Id_seq))

    db.commit()


def find_ID_Taxa(db_cursor, NCBI_TaxID, genotype, subtype, Id_genotype, Id_subtype ):
    #taxa = subtype if subtype else genotype
    NCBI_Taxa_ID = None
    genotype_Taxa_ID = None
    subtype_Taxa_ID = None

    if NCBI_TaxID:
        db_cursor.execute("SELECT Id_taxa FROM taxa_names WHERE Name=?", (NCBI_TaxID,))
        NCBI_Taxa_ID= db_cursor.fetchone()    # unique ??
        if not NCBI_Taxa_ID:
            print('No NCBI_Id found: NCBI_TaxID, genotype, subtype -> ', NCBI_TaxID, genotype, subtype)
            NCBI_Taxa_ID = None
        elif len(NCBI_Taxa_ID)!=1:
            print('Multiple NCBI_Id found: NCBI_Taxa_ID, NCBI_TaxID, genotype, subtype -> ', NCBI_Taxa_ID, NCBI_TaxID, genotype, subtype)
            NCBI_Taxa_ID = None
        else:
            NCBI_Taxa_ID = NCBI_Taxa_ID[0]

    if genotype:
        if NCBI_Taxa_ID:
           db_cursor.execute("SELECT Id_taxa FROM taxa_names JOIN taxa USING(Id_taxa) JOIN taxa_parents USING(Id_taxa) "
                             "WHERE taxa_names.Name=? AND taxa.Id_rank=? AND taxa_parents.parent=?",
                             (           genotype,          Id_genotype,        NCBI_Taxa_ID)        )
        else:
            db_cursor.execute("SELECT Id_taxa FROM taxa_names JOIN taxa USING(Id_taxa)   "
                              "WHERE taxa_names.Name=? AND taxa.Id_rank=?  ",
                                  (            genotype,          Id_genotype ))
        genotype_Taxa_ID = db_cursor.fetchone()    # unique ??
        if not genotype_Taxa_ID:
            if not subtype:
                #print('Swap genotype to subtype: NCBI_TaxID, genotype, subtype -> ', NCBI_TaxID, genotype, subtype)
                return find_ID_Taxa(db_cursor, NCBI_TaxID, None, genotype, Id_genotype, Id_subtype )
            print('Genotype not found: NCBI_TaxID, genotype, subtype -> ', NCBI_TaxID, genotype, subtype)
            genotype_Taxa_ID = None
        elif len(genotype_Taxa_ID)!=1:
            print('Multiple genotype found: genotype_Taxa_ID, NCBI_TaxID, genotype, subtype -> ',
                                            genotype_Taxa_ID, NCBI_TaxID, genotype, subtype)
            genotype_Taxa_ID = None
        else:
            genotype_Taxa_ID = genotype_Taxa_ID[0]

    if subtype:
        if genotype_Taxa_ID:
            if NCBI_Taxa_ID:
                db_cursor.execute(
                    "SELECT Id_taxa FROM taxa_names "
                    "               JOIN taxa               USING(Id_taxa) "
                    "               JOIN taxa_parents AS pg USING(Id_taxa) "
                    "               JOIN taxa_parents AS pn USING(Id_taxa) "
                   "WHERE taxa_names.Name=? "
                    " AND    taxa.Id_rank=? "
                    " AND       pg.parent=?"
                    " AND       pn.parent=?",
                    (subtype, Id_subtype, genotype_Taxa_ID, NCBI_Taxa_ID))
            else:
                db_cursor.execute(
                    "SELECT Id_taxa FROM taxa_names "
                    "               JOIN taxa               USING(Id_taxa) "
                    "               JOIN taxa_parents AS pg USING(Id_taxa) "
                   "WHERE taxa_names.Name=? "
                    " AND    taxa.Id_rank=? "
                    " AND       pg.parent=?",
                    (subtype, Id_subtype, genotype_Taxa_ID))
        else:
            if NCBI_Taxa_ID:
                db_cursor.execute(
                    "SELECT Id_taxa FROM taxa_names "
                    "               JOIN taxa               USING(Id_taxa) "
                    "               JOIN taxa_parents AS pn USING(Id_taxa) "
                   "WHERE taxa_names.Name=? "
                    " AND    taxa.Id_rank=? "
                    " AND       pn.parent=?",
                    (subtype, Id_subtype,  NCBI_Taxa_ID))
            else:
                db_cursor.execute(
                    "SELECT Id_taxa FROM taxa_names "
                    "               JOIN taxa               USING(Id_taxa) "
                   "WHERE taxa_names.Name=? "
                    " AND    taxa.Id_rank=? ",
                    (subtype, Id_subtype))

        subtype_Taxa_ID = db_cursor.fetchone()  # unique ??
        if not subtype_Taxa_ID:
                print('No subtype found: NCBI_TaxID, genotype, subtype -> ', NCBI_TaxID, genotype, subtype)
                subtype_Taxa_ID = None
        elif len(subtype_Taxa_ID) != 1:
            print('Multiple subtype found: subtype_Taxa_ID, NCBI_TaxID, genotype, subtype -> ',
                  subtype_Taxa_ID, NCBI_TaxID, genotype, subtype)
            subtype_Taxa_ID = None
        else:
            subtype_Taxa_ID = subtype_Taxa_ID[0]

    if subtype_Taxa_ID :
        return subtype_Taxa_ID

    if genotype_Taxa_ID:
        return genotype_Taxa_ID
    elif genotype:
        return find_ID_Taxa(db_cursor, NCBI_TaxID, None, genotype, Id_genotype, Id_subtype )

    if NCBI_Taxa_ID    : return NCBI_Taxa_ID


def reuse_GBdefinition_to_find_strain_isolate(definition, isolate, strain):
    if 'isolate' in definition:
        iso = definition.split('isolate')[1].split(',')[0]
        if iso[0] == ':':
            iso = iso[1:].strip()
        iso = iso.split()[0].strip()
        if iso[-1] == '.':
            iso = iso[:-1].strip()
        if isolate == '':
            isolate = iso
        else:
            if isolate != iso:
                isolate = isolate + ' or ' + iso
    if 'strain' in definition:
        st = definition.split('strain')[1].split(',')[0]
        if st[0] == ':':
            st = st[1:].strip()
        st = st.split()[0].strip()
        if st[-1] == '.':
            st = st[:-1].strip()
        if not strain:
            strain = st
        else:
            if strain != st:
                strain = strain + ' or ' + st
    return isolate, strain


def parse_GB_note(note):
    # print('Note=', note)
    genotype, subtype = None, None
    for n in note.split(';'):
        m = n.split()  #
        if len(m) == 1:
            m = m[0].split(':')
            if len(m) == 1:
                m = m[0].split('-')
        if m[0].lower().startswith                        ('genotype'):
            genotype = m[1].strip()
        elif m[0].lower().startswith                      ('subtype'):
            subtype = m[1].strip()
    return genotype, subtype


def parse_date(date :str):
    L=len(date)

    try:

        if L == 11 :                                    # 02-Mar-2010
            d=datetime.datetime.strptime(date, '%d-%b-%Y')
            return d.year, d.month, d.day

        elif L == 4 :                                   # 2011
            return date, None, None

        elif L == 8 :                                   # Sep-2011
                d=datetime.datetime.strptime(date, '%b-%Y')
                return d.year, d.month, None

        elif L == 9:
            d=date.split('/')[0]                        # 2013/2014
            return d, None, None

        elif L == 10 :                                  # 2016-05-12
            d=datetime.datetime.strptime(date, '%Y-%m-%d')
            return d.year, d.month, d.day

        elif L == 7 :                                   # 2016-05
            d=datetime.datetime.strptime(date, '%Y-%m')
            return d.year, d.month, None

        elif L == 3 :                                   # Sep       ???????
            d=datetime.datetime.strptime(date, '%b')
            return None, d.month, None

        else:
            raise ValueError

    except ValueError:
        print('Unable to parse date: ', date, ' of length ', len(date) )
        return None, None, None

def parse_source(host:str, source:str):
    s = None
    h = None
    o = None
    if host:
        o= host.lower()
        if source:
            source: source = source.lower()
            o += source
    else:
        if source:
            source: source = source.lower()
            o = source
        else:  return None, None



    if   any(syn in o for syn in ['huma', 'homo sapiens', 'patient', 'donor', 'clinical' ]):
        h = 'human'
    elif any(syn in o for syn in ['boar', 'sus scrofa']):    # wild boar
        if 'sus scrofa domesticus' in o:
            h = 'swine'
        else:
            h = 'wild boar'
    elif any(syn in o for syn in ['swine', 'pig', 'porcine', 'pork', 'sus scrofa domesticus', 'figatellu']):
        h = 'swine'
    elif any(syn in o for syn in ['rattus', 'rat']):   # Rattus flavipectus?? =Rattus tanezumi
        h = 'rat'
    elif any(syn in o for syn in ['bandicota indica']):   # greater bandicoot rat (Bandicota indica) is a species of rodent in the family Muridae found in Bangladesh, China,
        h = 'Bandicota indica'
    elif any(syn in o for syn in ['goat']):      # cabra, Capra aegagrus hircus
        h = 'goat'
    elif any(syn in o for syn in ['ferret', 'mustela putorius furo']):
        h = 'ferret'
    elif any(syn in o for syn in ['mustela putorius']):
        h = 'polecat'
    elif any(syn in o for syn in ['Vulpes vulpes', 'fox']):  #  red fox (Vulpes vulpes), largest of the true foxes,
        h = 'fox'
    elif any(syn in o for syn in ['breadcrumb sponge', 'halichondria panicea']):
        h = 'sponge'
    elif any(syn in o for syn in ['treeshrew', 'tree shrew']):
        h = 'treeshrew'
    elif any(syn in o for syn in ['shrew', 'suncus murinus']): #  Asian house shrew (Suncus murinus) grey musk shrew, Asian musk shrew, or money shrew is a widespread, adaptable species of shrew
        h = 'shrew'
    elif any(syn in o for syn in ['oryctolagus cuniculus', 'rabbit']):
        h = 'rabbit'
    elif any(syn in o for syn in ['chicken', 'hens']):
        h = 'chicken'
    elif any(syn in o for syn in ['gallus gallus']):  #  red junglefowl (Gallus gallus) is a tropical member of the family Phasianidae. It is the primary progenitor of the domestic chicken
        h = 'Gallus gallus'
    elif any(syn in o for syn in ['monkey', 'Macaca mulatta']):
        h = 'monkey'
    elif any(syn in o for syn in ['cattle', 'cow']):
        h = 'cattle'
    elif any(syn in o for syn in ['tiger']):
        h = 'tiger'
    elif any(syn in o for syn in ['roe deer']):       # venado, Capreolus capreolus), also known as the western roe deer, chevreuil, or simply roe deer or roe
        h = 'roe deer'
    elif any(syn in o for syn in ['moose', 'alces alces']):  #  moose (North America) or elk (Eurasia), Alces alces, is the largest extant species in the deer family
        h = 'moose'
    elif any(syn in o for syn in ['red deer']):       # venado, red deer (Cervus elaphus)
        h = 'red deer'
    elif any(syn in o for syn in ['deer']):  # venado,
        h = 'deer'
    elif any(syn in o for syn in ['camelus']):        # Bactrian camel
        h = 'camel'
    elif any(syn in o for syn in ['thrush']):         # Turdidae, of passerine birds
        h = 'thrush'
    elif any(syn in o for syn in ['feral pigeon']):       # palomas, feral pigeons (Columba livia domestica), city doves, city pigeons, street pigeons, or flying rats
        h = 'feral pigeon'
    elif any(syn in o for syn in ['buzzard']):       # buzzard is the common name of several species of bird of prey:
        h = 'buzzard'
    elif any(syn in o for syn in ['owl']):       # Owls are birds from the order Strigiformes
        h = 'owl'
    elif any(syn in o for syn in ['sheep', 'ovine']):       #  sheep (Ovis aries)
        h = 'sheep'
    elif any(syn in o for syn in ['falco']):       #   common kestrel (Falco tinnunculus) is a bird of prey species belonging to the kestrel group of the falcon family Falconidae.
        h = 'falcon'
    elif any(syn in o for syn in ['mystacina tuberculata',
                                  'vampyrodes caraccioli',
                                  'rhinolophus ferrumequinum',
                                  'myotis',
                                  'eptesicus serotinus',
                                  'hipposideros abae']):       # serotine bat (Eptesicus serotinus), also known as the common serotine bat, big brown bat or silky bat --   lesser short-tailed bat (Mystacina tuberculata) – pekapeka-tou-poto in Māori – is the only living species of bat in the family Mystacinidae,[1] and is endemic to New Zealand.
        h = 'bat'
    elif any(syn in o for syn in ['arctocephalus australis']):       #   South American fur seal (Arctocephalus australis)
        h = 'seal'
    elif any(syn in o for syn in ['vulture']):       #   Himalayan vulture or Himalayan griffon vulture (Gyps himalayensis) is an Old World vulture in the family Accipitridae.
        h = 'vulture'
    elif any(syn in o for syn in ['dolphin']):       #
        h = 'dolphin'
    elif any(syn in o for syn in ['bos grunniens']):       #
        h = 'yak'
    elif any(syn in o for syn in ['mink', 'neovison vison']):       #
        h = 'mink'
    elif any(syn in o for syn in ['egret']):       # little egret (Egretta garzetta)
        h = 'egret'
    elif any(syn in o for syn in ['ruditapes philippinarum', 'scapharca subcrenata', 'anadara granosa']):       #
        h = 'clam'
    elif any(syn in o for syn in ['mytilus galloprovincialis']):  #
        h = 'mussel'
    elif any(syn in o for syn in ['oyster']):  #
        h = 'oyster'
    elif any(syn in o for syn in ['Felis catus']):       #
        h = 'cat'
    elif any(syn in o for syn in ['trout']):       # brook trout, cutthroat brook
        h = 'trout'
    elif any(syn in o for syn in ['herpestes javanicus', 'mongoose']):       #
        h = 'mongoose'
    elif any(syn in o for syn in ['horse']):       #
        h = 'horse'
    elif any(syn in o for syn in ['leopard']):       #
        h = 'leopard'
    elif any(syn in o for syn in ['bear']):       #
        h = 'bear'
    elif any(syn in o for syn in ['crowned crane']):       #
        h = 'crowned crane'
    elif any(syn in o for syn in ['pheasant']):       #
        h = 'pheasant'




    if not source: return h, None



    if   any(syn in source for syn in ['sera', 'serum']):
        s = 'sera'
    elif any(syn in source for syn in ['plasma']):
        s = 'plasma'
    elif any(syn in source for syn in ['liver', 'figatellu']):
        s = 'liver'
    elif any(syn in source for syn in ['feces',
                                       'anal swab',
                                       'manure',
                                       'stool',
                                       'fecal',
                                       'intestinal content',
                                       'caecal contents',
                                       'fecces']):
        s = 'feces'
    elif any(syn in source for syn in ['bile', 'gall']):
        s = 'bile'
    elif any(syn in source for syn in ['milk']):
        s = 'milk'
    elif any(syn in source for syn in ['blood']):
        s = 'blood'
    elif any(syn in source for syn in ['brain']):
        s = 'brain'
    elif any(syn in source for syn in ['heart']):
        s = 'heart'
    elif any(syn in source for syn in ['kidney']):
        s = 'kidney'
    elif any(syn in source for syn in ['spleen']):
        s = 'spleen'
    elif any(syn in source for syn in ['urine']):
        s = 'urine'
    elif any(syn in source for syn in ['nares']):
        s = 'nares'
    elif any(syn in source for syn in ['colon']):
        s = 'colon'
    elif any(syn in source for syn in ['egg']):
        s = 'egg'
    elif any(syn in source for syn in ['sausage']):
        s = 'sausage'
    elif any(syn in source for syn in ['flesh', 'pork', 'sausage', 'muscle', 'meat']):
        s = 'muscle'
    elif any(syn in source for syn in ['cerebrospinal fluid', 'csf']):
        s = 'CSF'
    elif any(syn in source for syn in ['sewage', 'effluent', 'slurry']):   #   ????? feces ??
        s = 'sewage'
    elif any(syn in source for syn in ['river']):   #   ?????
        s = 'river water'
    elif any(syn in source for syn in ['culture supernatant', 'cell culture', 'hepatocytes']):   #   ?????
        s = 'cell culture'
    elif any(syn in source for syn in ['pool']):       #   ?????
        s = 'pool'
    elif any(syn in source for syn in ['soil']):  # ?????
        s = 'soil'
    elif any(syn in source for syn in ['swab']):       #   ?????
        s = 'swab'
    elif any(syn in source for syn in ['wastewater', 'slaughterhouse']):       #   ?????
        s = 'wastewater'
    elif any(syn in source for syn in ['water']):       #   ?????
        s = 'water'
    elif any(syn in source for syn in ['strawberry']):       #   ?????
        s = 'strawberry'

    return h, s

def create_all( ):

    newly=True   # False
    country=True # False

    print('Creating db BioSQL...')

#    dbBioSQL=createBioSQL(newly)



    print('Creating db...')
    sdb = create(newly)

    if newly or country:
        read_country_codes(sdb)

    ref_name = "M73218"
    if newly:
        print('Parsing GB_flat_file...')
        # parseGB(sdb, r'C:/Prog/HEV/data/temp/HEV-g3marked_all_2017-09-27.  577 seq.sequence.gb')
        parseGB(sdb, r'../data/temp/HEV_all.sequence.gb')
        print('Parsing the big alignment...')
        ID_align, ref = parse_full_fasta_Align(sdb, ref_name,'../alignment/HEV.fas')
        print(ref)
    else:
        print('Cleaning parsed Excel...')
        clean_parsed_Excel(sdb)
        ID_align=1      # ??

    print('Calcule reference positions...')
    ref = ref_pos(sdb, ID_align, ref_name) # , ref_name
    print(ref)

    print('Parse the Excel file table...')
    parse_HEV_xlsm(sdb, '../data/temp/HEVsubtypingMEGAut - Kopie.xlsm')

    print('Done !!!')
    sdb.close()

    # """


if __name__ == '__main__':
    create_all( )

# exit(0)
# """
